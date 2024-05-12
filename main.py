import json#jsonファイル作成
import openpyxl#excelファイル連携
import glob #指定フォルダから画像抽出 


##################################################
############　　　　以下、使い方　　　　###############
##################################################

#【手順１】取り込むexcelファイルを「main.py」と同じディレクトリに入れて、以下でファイル名を指定（''の中を変更）
excelFile = 'forTest.xlsx'

#【手順２】取り込むexcelファイルの初めの行と終わりの行を指定
startRow = 3
endRow = 5

#【手順３】プログラムを走らせる

##################################################
#############　　　　以下、本体　　　　################
##################################################


#excelファイルをオープン
wb = openpyxl.load_workbook(excelFile)
#シートを選択
sheet = wb['Sheet1']


#対象エクセルの行数分だけ繰り返し
count = startRow
while count < endRow +1:

#imageファイルをフォルダから抽出
    imageFromExcel = sheet.cell(row=count, column=2).value
    img = glob.glob('/Users/yokot/Documents/vscode_Json_LLAVA/Json_Creator'+ imageFromExcel)

#promptを生成（最後の一文は「この犬をペットにしたいですか？」）
    margeCell = sheet.cell(row=count, column=5).value\
                +'竣工の'+sheet.cell(row=count, column=6).value\
                    +'の'+sheet.cell(row=count, column=7).value\
                        +'の写真です。'\
                            +'この橋梁は精密な検査を必要としますか？'   

#JSONファイルを作成
    tuningSet = {
        sheet.cell(row=count, column=1).value:{
        'image':img,
        'prompt':margeCell,
        'additionalInfo':{'fixedInfo':{
                              'item1':sheet.cell(row=count, column=8).value,},
                          'not_fixedInfo':{
                              'item2':sheet.cell(row=count, column=9).value,
                              'item3':sheet.cell(row=count, column=10).value,
                              'item4':sheet.cell(row=count, column=11).value}},
        'answer':sheet.cell(row=count, column=12).value\
                     + 'その理由は次のとおりです'+sheet.cell(row=count, column=13).value\
        }
    }

    out = json.dumps(tuningSet, indent=4, ensure_ascii=False)
    with open('LLava.json2', 'a', encoding='utf-8') as fout:
        fout.write(out)

    count += 1


else:
    print('☆☆☆☆☆jsonファイル出力完了!おめでとう！！☆☆☆☆☆')




##################################################
#############　　　　以下、材料　　　　################
##################################################

#def print_hi(name):
#    # Use a breakpoint in the code line below to debug your script.
#    print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.

# Press the green button in the gutter to run the script.
#if __name__ == '__main__':
#    print_hi('PyCharm')


#testSet = {
#    'test1':{
#        'hours':'7:30',
#        'menus':['pan','bacon']
#    },
#
#    'test2': {
#        '時間': '7:30',
#        'メニュー': ['トースト', 'ベーコン']
#    }
#}

#画像を表示（デバック用）
#image = cv2.imread('/Users/tsurutashuichi/PycharmProjects/pythonProject2/image_strage/dog.犬.1-1.jpg')
#cv2.imshow('image_test',image)
#cv2.waitKey(0)
#cv2.destroyWindow('image_test')
#print(cv2.__version__)

#セルの値を取得
#for i in range(2,8):
#    prompt = sheet.cell(row=i, column=1).value \
#                +'と' \
#                    + sheet.cell(row=i, column=2).value
#    print(prompt)

#JSONファイルに追加する辞書データをupdateで追加する
#    with open('sample.json','r')as file:
#        data = json.load(file)
#        data.update(add_data)

#JSONファイルにデータを追加(write)する
#    with open('sample.json', 'a')as file:
#        json.dump(data, file, indent=2, ensure_ascii=False)


#imageファイルをフォルダから抽出
#imageFileName = "dog.犬.1-1.jpg"
#imagePath = glob.glob("/Users/tsurutashuichi/PycharmProjects/pythonProject2/image_strage/"+ imageFileName)
#print(imageFileName)
#print(imagePath)